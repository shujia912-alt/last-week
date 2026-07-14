#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理「正式-上周派单情况.xlsx」：
1. 在 DEVELOP_ORDER 工作表中，将“需求确认时间”列转换为仅保留日期（yyyy-mm-dd）
2. 删除满足条件的整行
3. 将 DEVELOP_ORDER 数据透视到新建 sheet1
4. 将 sheet1 从第5行开始到最后一行复制到新建工作表“派单情况”
5. 删除供应商名称中包含指定关键字的行
6. 新增“纸样日产能”列
7. 新增“纸样周产能”列
8. 根据“供应商产能”表匹配纸样日产能
9. 计算纸样周产能
10. 汇总总计行
"""

from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FILE_PATH = Path.home() / "Desktop" / "正式-上周派单情况.xlsx"


def normalize_date_only(value):
    """将日期时间转换为 yyyy-mm-dd，只保留日期部分。"""
    if pd.isna(value) or value == "":
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return value
        return ts.date()
    except Exception:
        return value


def main():
    wb = load_workbook(FILE_PATH)

    # =========================
    # 1&2. 处理 DEVELOP_ORDER
    # =========================
    if "DEVELOP_ORDER" not in wb.sheetnames:
        raise ValueError("未找到工作表：DEVELOP_ORDER")

    ws = wb["DEVELOP_ORDER"]

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    header_map = {name: idx + 1 for idx, name in enumerate(headers) if name is not None}

    required_cols = ["需求确认时间", "开发单状态", "开发跟单签收时间", "复色场景", "供应商", "SPU"]
    for col in required_cols:
        if col not in header_map:
            raise ValueError(f"DEVELOP_ORDER 中缺少列：{col}")

    col_req_time = header_map["需求确认时间"]
    col_status = header_map["开发单状态"]
    col_sign_time = header_map["开发跟单签收时间"]

    # 1. “需求确认时间”仅保留日期
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_req_time)
        new_val = normalize_date_only(cell.value)
        cell.value = new_val
        if isinstance(new_val, datetime):
            cell.number_format = "yyyy-mm-dd"
        else:
            cell.number_format = "yyyy-mm-dd"

    # 2. 删除满足条件的整行
    rows_to_delete = []
    for row in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row, column=col_status).value
        sign_val = ws.cell(row=row, column=col_sign_time).value
        if str(status_val).strip() == "已关闭" and (sign_val is None or str(sign_val).strip() == ""):
            rows_to_delete.append(row)

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)

    # 重新读取数据，准备透视
    data = ws.values
    df = pd.DataFrame(data)
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)

    # 清理字段
    for col in ["需求确认时间", "复色场景", "供应商", "SPU"]:
        if col in df.columns:
            df[col] = df[col].astype(str).replace("nan", "").str.strip()

    # “需求确认时间”转日期字符串
    if "需求确认时间" in df.columns:
        df["需求确认时间"] = pd.to_datetime(df["需求确认时间"], errors="coerce").dt.strftime("%Y-%m-%d")

    # 3. 透视到新建 sheet1
    # 筛选：复色场景 去掉 “爆款复色”“普通加色”
    if "复色场景" in df.columns:
        df = df[~df["复色场景"].isin(["爆款复色", "普通加色"])]

    # 筛去“需求确认时间”空白
    df = df[df["需求确认时间"].notna() & (df["需求确认时间"].astype(str).str.strip() != "")]

    # pivot
    pivot = pd.pivot_table(
        df,
        index=["供应商"],
        columns=["需求确认时间"],
        values="SPU",
        aggfunc="count",
        fill_value=0,
        margins=True,
        margins_name="总计"
    )

    # 保证供应商落到 A5
    sheet1_name = "sheet1"
    if sheet1_name in wb.sheetnames:
        del wb[sheet1_name]
    ws1 = wb.create_sheet(sheet1_name)

    # 写入透视表，从 A5 开始
    start_row = 5
    start_col = 1

    # 写列头
    ws1.cell(row=start_row, column=start_col, value="供应商")
    for j, col_name in enumerate(pivot.columns, start=start_col + 1):
        ws1.cell(row=start_row, column=j, value=col_name)

    # 写数据
    for i, idx_name in enumerate(pivot.index, start=start_row + 1):
        ws1.cell(row=i, column=start_col, value=idx_name)
        for j, val in enumerate(pivot.loc[idx_name].tolist(), start=start_col + 1):
            ws1.cell(row=i, column=j, value=val)

    # 4. 将 sheet1 从第5行开始到最后一行复制到“派单情况”
    target_name = "派单情况"
    if target_name in wb.sheetnames:
        del wb[target_name]
    ws2 = wb.create_sheet(target_name)

    max_row_1 = ws1.max_row
    max_col_1 = ws1.max_column

    for r in range(5, max_row_1 + 1):
        for c in range(1, max_col_1 + 1):
            ws2.cell(row=r - 4, column=c, value=ws1.cell(row=r, column=c).value)

    # 读取“派单情况”到 DataFrame
    data2 = ws2.values
    df2 = pd.DataFrame(data2)
    df2.columns = df2.iloc[0]
    df2 = df2.iloc[1:].reset_index(drop=True)

    # 去掉空列名
    df2 = df2.loc[:, df2.columns.notna()]

    # 5. 删除供应商名称包含关键字的行
    keywords = ["毛衣", "泳衣", "牛仔"]
    if "供应商" not in df2.columns:
        raise ValueError("派单情况中未找到“供应商”列")

    def should_drop_supplier(x):
        x = "" if pd.isna(x) else str(x)
        return any(k in x for k in keywords)

    df2 = df2[~df2["供应商"].apply(should_drop_supplier)].reset_index(drop=True)

    # 6. 在“供应商”列右侧新增“纸样日产能”
    supplier_idx = df2.columns.get_loc("供应商")
    df2.insert(supplier_idx + 1, "纸样日产能", 0)

    # 7. 在“总计”列右侧新增“纸样周产能”
    if "总计" not in df2.columns:
        raise ValueError("派单情况中未找到“总计”列")
    total_idx = df2.columns.get_loc("总计")
    df2.insert(total_idx + 1, "纸样周产能", 0)

    # 8. 匹配“供应商产能”表，填入纸样日产能
    if "供应商产能" not in wb.sheetnames:
        raise ValueError("未找到工作表：供应商产能")

    cap_ws = wb["供应商产能"]
    cap_data = cap_ws.values
    cap_df = pd.DataFrame(cap_data)
    cap_df.columns = cap_df.iloc[0]
    cap_df = cap_df.iloc[1:].reset_index(drop=True)

    # 供应商产能：A列供应商，B列纸样日产能，C列车版日产能
    cap_df = cap_df.iloc[:, :3]
    cap_df.columns = ["供应商", "纸样日产能", "车版日产能"]
    cap_df["供应商"] = cap_df["供应商"].astype(str).str.strip()

    cap_map = {}
    for _, row in cap_df.iterrows():
        supplier = str(row["供应商"]).strip()
        try:
            capacity = float(row["纸样日产能"])
            if capacity.is_integer():
                capacity = int(capacity)
        except Exception:
            capacity = 0
        cap_map[supplier] = capacity

    def get_daily_capacity(supplier):
        if pd.isna(supplier):
            return 0
        supplier = str(supplier).strip()
        return cap_map.get(supplier, 0)

    df2["纸样日产能"] = df2["供应商"].apply(get_daily_capacity)

    # 9. 计算纸样周产能
    def calc_weekly_capacity(row):
        supplier = "" if pd.isna(row["供应商"]) else str(row["供应商"]).strip()
        daily = row["纸样日产能"]
        if supplier == "cider内部板房":
            return daily * 5
        return daily * 6

    df2["纸样周产能"] = df2.apply(calc_weekly_capacity, axis=1)

    # 10. 最后“总计”行求和
    # pivot_table(margins=True) 已经生成了“总计”行，这里只补充新增列的汇总值，避免重复生成第二个总计行
    if "总计" in df2["供应商"].astype(str).values:
        total_row_idx = df2.index[df2["供应商"].astype(str) == "总计"][0]
        df2.loc[total_row_idx, "纸样日产能"] = pd.to_numeric(df2["纸样日产能"], errors="coerce").fillna(0).sum()
        df2.loc[total_row_idx, "纸样周产能"] = pd.to_numeric(df2["纸样周产能"], errors="coerce").fillna(0).sum()
    else:
        sum_row = {col: "" for col in df2.columns}
        sum_row["供应商"] = "总计"
        sum_row["纸样日产能"] = pd.to_numeric(df2["纸样日产能"], errors="coerce").fillna(0).sum()
        sum_row["纸样周产能"] = pd.to_numeric(df2["纸样周产能"], errors="coerce").fillna(0).sum()
        if "总计" in df2.columns:
            sum_row["总计"] = pd.to_numeric(df2["总计"], errors="coerce").fillna(0).sum()
        df2 = pd.concat([df2, pd.DataFrame([sum_row])], ignore_index=True)

    # 写回“派单情况”
    # 先清空旧内容
    ws2.delete_rows(1, ws2.max_row)

    for r_idx, row in enumerate([df2.columns.tolist()] + df2.values.tolist(), start=1):
        for c_idx, val in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=val)

    # 调整行高，让表格更适配
    for r in range(1, ws2.max_row + 1):
        ws2.row_dimensions[r].height = 20

    # 列宽简单优化
    for col_idx in range(1, ws2.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, ws2.max_row + 1):
            v = ws2.cell(r, col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 25)

    # 保存
    wb.save(FILE_PATH)
    print(f"处理完成，已保存到：{FILE_PATH}")


if __name__ == "__main__":
    main()